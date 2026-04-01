"""
Excel 转 JSON 转换工具（支持 .xls 和 .xlsx）
用法: python excel_to_json.py
自动转换 assets 文件夹下的所有 Excel 文件

支持格式:
- *.xlsx -> 使用 openpyxl
- *.xls  -> 使用 xlrd
"""

import csv
import json
import os
import openpyxl
import xlrd
from datetime import datetime
from pathlib import Path

# 配置路径
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ASSETS_DIR = os.path.join(PROJECT_ROOT, "assets")

# 各文件的字段映射配置
CONFIGS = {
    "cardInfo": {
        "field_map": {
            "base_cardName": "id",
            "index": "index",
            "base_displayName": "displayName",
            "base_cardClass": "cardClass",
            "base_desc": "desc",
            "base_price": "price",
            "base_card": "card",
            "base_max": "max",
            "site_area": "siteArea",
            "npc_sched": "npcSched",
            "food_HP": "foodHP",
            "eventId": "eventId"
        },
        "int_fields": {"index", "price", "card", "max", "siteArea", "npcSched", "foodHP", "eventId"},
        "key_field": "id"
    },
    "eventInfo": {
        "field_map": {
            "eventId": "eventId",
            "title": "title",
            "condition": "condition",
            "card_1": "card_1",
            "card_2": "card_2",
            "card_3": "card_3",
            "card_4": "card_4",
            "card_5": "card_5",
            "card_6": "card_6",
            "result": "result",
            "description_begain": "description_begain",
            "description_result": "description_result"
        },
        "int_fields": {"eventId"},  # card_1 到 card_6 是字符串，不是整数
        "key_field": "eventId"
    }
}


def read_excel_xlsx(file_path: str):
    """读取 .xlsx 文件"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    items = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        items[row[0]] = row_dict
    
    wb.close()
    return items


def read_excel_xls(file_path: str):
    """读取 .xls 文件"""
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)
    headers = ws.row_values(0)
    
    items = {}
    for row_idx in range(1, ws.nrows):
        row_values = ws.row_values(row_idx)
        if not row_values[0]:
            continue
        row_dict = {headers[i]: row_values[i] for i in range(len(headers))}
        items[row_values[0]] = row_dict
    
    return items


def convert_excel_file(excel_name: str, config: dict) -> dict:
    """转换单个 Excel 文件"""
    ext = ".xlsx" if os.path.exists(os.path.join(ASSETS_DIR, f"{excel_name}.xlsx")) else ".xls"
    excel_path = os.path.join(ASSETS_DIR, f"{excel_name}{ext}")
    
    print(f"\n正在处理: {excel_name}{ext}")
    
    if ext == ".xlsx":
        raw_data = read_excel_xlsx(excel_path)
    else:
        raw_data = read_excel_xls(excel_path)
    
    # 转换字段名和类型
    items = {}
    for key, row in raw_data.items():
        # 处理 key 类型（可能是 float）
        if isinstance(key, float):
            key = int(key)
        
        item = {}
        for excel_field, json_field in config["field_map"].items():
            value = row.get(excel_field, "")
            
            # 整数转换
            if json_field in config["int_fields"]:
                try:
                    if isinstance(value, float):
                        value = int(value)
                    else:
                        value = int(value) if value else 0
                except (ValueError, TypeError):
                    value = 0
            
            item[json_field] = value
        
        items[key] = item
    
    print(f"  转换完成: {len(items)} 条记录")
    return items


def main():
    print("=" * 50)
    print("Excel 转 JSON 工具")
    print("=" * 50)
    
    all_data = {}
    
    for excel_name, config in CONFIGS.items():
        excel_path_xlsx = os.path.join(ASSETS_DIR, f"{excel_name}.xlsx")
        excel_path_xls = os.path.join(ASSETS_DIR, f"{excel_name}.xls")
        
        if os.path.exists(excel_path_xlsx):
            items = convert_excel_file(excel_name, config)
            all_data[excel_name] = items
        elif os.path.exists(excel_path_xls):
            items = convert_excel_file(excel_name, config)
            all_data[excel_name] = items
        else:
            print(f"  跳过: {excel_name} (文件不存在)")
    
    # 输出 cardData.json
    output_path = os.path.join(ASSETS_DIR, "cardData.json")
    output = {
        "_metadata": {
            "version": "1.0",
            "description": "卡牌游戏配置数据",
            "generated": datetime.now().isoformat()
        },
        "cardInfo": all_data.get("cardInfo", {}),
        "eventInfo": all_data.get("eventInfo", {})
    }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print("\n" + "=" * 50)
    print(f"全部转换完成!")
    print(f"输出文件: {output_path}")
    print("=" * 50)


if __name__ == "__main__":
    main()
